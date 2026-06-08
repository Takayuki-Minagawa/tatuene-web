#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel原本から「ワークブックモデル」を抽出する。
数式は一切書き換えず、原本の文字列をそのまま保存する（契約・東京都評価との同一性担保）。

出力:
  src/data/workbook-model.json  ... 全シートの数式/値 + キャッシュ計算値 + 結合 + 寸法
  src/data/material-master.json ... 建材マスタ(ドロップダウン候補)
  public/assets/                ... ロゴ等の画像(png/jpeg)
"""
import json, os, sys, re, shutil, zipfile
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ---- テーマ色 解決 ----
_THEME_CACHE = {}

def load_theme_colors(xlsx):
    """theme1.xml の clrScheme を XML順で返す（hex 6桁）。"""
    if "colors" in _THEME_CACHE:
        return _THEME_CACHE["colors"]
    colors = []
    try:
        with zipfile.ZipFile(xlsx) as z:
            t = z.read("xl/theme/theme1.xml").decode("utf-8")
        scheme = re.search(r"<a:clrScheme.*?</a:clrScheme>", t, re.S).group(0)
        # 各色要素から srgbClr val / sysClr lastClr を取得（XML出現順）
        for m in re.finditer(r"<a:(dk1|lt1|dk2|lt2|accent1|accent2|accent3|accent4|accent5|accent6|hlink|folHlink)>(.*?)</a:\1>", scheme, re.S):
            body = m.group(2)
            srgb = re.search(r'srgbClr val="([0-9A-Fa-f]{6})"', body)
            sysc = re.search(r'sysClr[^>]*lastClr="([0-9A-Fa-f]{6})"', body)
            colors.append((srgb or sysc).group(1).upper() if (srgb or sysc) else "FFFFFF")
    except Exception:
        colors = ["000000", "FFFFFF", "44546A", "E7E6E6"] + ["4472C4"] * 8
    _THEME_CACHE["colors"] = colors
    return colors


def apply_tint(hexrgb, tint):
    r = int(hexrgb[0:2], 16); g = int(hexrgb[2:4], 16); b = int(hexrgb[4:6], 16)
    def adj(c):
        if tint is None or tint == 0:
            return c
        if tint < 0:
            return int(round(c * (1 + tint)))
        return int(round(c * (1 - tint) + 255 * tint))
    return f"{adj(r):02X}{adj(g):02X}{adj(b):02X}"


def resolve_color(color, xlsx):
    """openpyxl Color → '#RRGGBB' or None"""
    if color is None:
        return None
    try:
        t = color.type
        if t == "rgb" and isinstance(color.rgb, str) and len(color.rgb) == 8:
            return "#" + color.rgb[2:].upper()
        if t == "theme":
            theme = load_theme_colors(xlsx)
            idx = color.theme
            # Excel表示順: 0,1 は背景/文字でXMLのdk1/lt1と入れ替わる
            order = [1, 0, 3, 2, 4, 5, 6, 7, 8, 9, 10, 11]
            if idx < len(order) and order[idx] < len(theme):
                base = theme[order[idx]]
                return "#" + apply_tint(base, getattr(color, "tint", 0) or 0)
        if t == "indexed":
            from openpyxl.styles.colors import COLOR_INDEX
            i = color.indexed
            if i is not None and i < len(COLOR_INDEX):
                v = COLOR_INDEX[i]
                if isinstance(v, str) and len(v) == 8:
                    return "#" + v[2:].upper()
    except Exception:
        return None
    return None

XLSX = os.path.join(ROOT, "..", "data", "Web版_かつエネ断熱シミュレーターVer1-7-6.xlsx")
OUT_MODEL = os.path.join(ROOT, "src", "data", "workbook-model.json")
OUT_MASTER = os.path.join(ROOT, "src", "data", "material-master.json")
ASSET_DIR = os.path.join(ROOT, "public", "assets")


def cell_content(c):
    """書込み用セル内容: 数式は '=...' 文字列、それ以外は生値。空は None。"""
    v = c.value
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("="):
        return v  # 原本の数式をそのまま
    return v


# 除外する画像（間取り図＝図面/キープランは別途作成のため不要）
EXCLUDE_IMAGES = {"image3.png"}
EMU_PER_PX = 9525  # 914400 EMU/inch ÷ 96 px/inch


def images_by_sheet():
    """各シートの drawing から PNG/JPEG 画像の配置（アンカー）を抽出。
    EMF（web非対応）と間取り図(image3)は除外。返り値: {シート名: [画像配置...]}"""
    import zipfile as _zip
    result = {}
    with _zip.ZipFile(XLSX) as z:
        names = set(z.namelist())
        relsd = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid2file = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', relsd))
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml):
            sheet_name, rid = m.group(1), m.group(2)
            sf = rid2file.get(rid)
            if not sf:
                continue
            wrels_path = f"xl/worksheets/_rels/{os.path.basename(sf)}.rels"
            if wrels_path not in names:
                continue
            dm = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', z.read(wrels_path).decode("utf-8"))
            if not dm:
                continue
            dfile = "xl/" + dm.group(1)
            dxml = z.read(dfile).decode("utf-8")
            drels_path = f"xl/drawings/_rels/{os.path.basename(dfile)}.rels"
            embed2media = {}
            if drels_path in names:
                embed2media = dict(re.findall(
                    r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"',
                    z.read(drels_path).decode("utf-8")))
            imgs = []
            for a in re.finditer(r"<xdr:twoCellAnchor.*?</xdr:twoCellAnchor>", dxml, re.S):
                block = a.group(0)
                fr = re.search(r"<xdr:from><xdr:col>(\d+)</xdr:col><xdr:colOff>(-?\d+)</xdr:colOff><xdr:row>(\d+)</xdr:row><xdr:rowOff>(-?\d+)", block)
                to = re.search(r"<xdr:to><xdr:col>(\d+)</xdr:col><xdr:colOff>(-?\d+)</xdr:colOff><xdr:row>(\d+)</xdr:row><xdr:rowOff>(-?\d+)", block)
                if not (fr and to):
                    continue
                for emb in re.findall(r'r:embed="(rId\d+)"', block):
                    media = embed2media.get(emb)
                    if not media:
                        continue
                    ext = media.rsplit(".", 1)[-1].lower()
                    if ext not in ("png", "jpg", "jpeg"):
                        continue  # EMF等は除外
                    if media in EXCLUDE_IMAGES:
                        continue  # 間取り図は除外
                    imgs.append({
                        "file": media,
                        "fromCol": int(fr.group(1)), "fromColOff": int(fr.group(2)),
                        "fromRow": int(fr.group(3)), "fromRowOff": int(fr.group(4)),
                        "toCol": int(to.group(1)), "toColOff": int(to.group(2)),
                        "toRow": int(to.group(3)), "toRowOff": int(to.group(4)),
                    })
            if imgs:
                result[sheet_name] = imgs
    return result


def drawing_slots_by_sheet():
    """各シートの drawing から EMF/WMF アンカー（＝図面の配置枠）を抽出。
    ユーザーが図面をアップロードして配置・注釈する枠として扱う。
    返り値: {シート名: [{id, fromCol,...,toRow,toRowOff}...]}（ドキュメント順）"""
    import zipfile as _zip
    result = {}
    with _zip.ZipFile(XLSX) as z:
        names = set(z.namelist())
        relsd = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid2file = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', relsd))
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml):
            sheet_name, rid = m.group(1), m.group(2)
            sf = rid2file.get(rid)
            if not sf:
                continue
            wrels_path = f"xl/worksheets/_rels/{os.path.basename(sf)}.rels"
            if wrels_path not in names:
                continue
            dm = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', z.read(wrels_path).decode("utf-8"))
            if not dm:
                continue
            dfile = "xl/" + dm.group(1)
            dxml = z.read(dfile).decode("utf-8")
            drels_path = f"xl/drawings/_rels/{os.path.basename(dfile)}.rels"
            embed2media = {}
            if drels_path in names:
                embed2media = dict(re.findall(
                    r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"',
                    z.read(drels_path).decode("utf-8")))
            slots = []
            for a in re.finditer(r"<xdr:twoCellAnchor.*?</xdr:twoCellAnchor>", dxml, re.S):
                block = a.group(0)
                fr = re.search(r"<xdr:from><xdr:col>(\d+)</xdr:col><xdr:colOff>(-?\d+)</xdr:colOff><xdr:row>(\d+)</xdr:row><xdr:rowOff>(-?\d+)", block)
                to = re.search(r"<xdr:to><xdr:col>(\d+)</xdr:col><xdr:colOff>(-?\d+)</xdr:colOff><xdr:row>(\d+)</xdr:row><xdr:rowOff>(-?\d+)", block)
                if not (fr and to):
                    continue
                is_emf = False
                for emb in re.findall(r'r:embed="(rId\d+)"', block):
                    media = embed2media.get(emb, "")
                    if media.rsplit(".", 1)[-1].lower() in ("emf", "wmf"):
                        is_emf = True
                if not is_emf:
                    continue
                slots.append({
                    "id": f"slot{len(slots) + 1}",
                    "fromCol": int(fr.group(1)), "fromColOff": int(fr.group(2)),
                    "fromRow": int(fr.group(3)), "fromRowOff": int(fr.group(4)),
                    "toCol": int(to.group(1)), "toColOff": int(to.group(2)),
                    "toRow": int(to.group(3)), "toRowOff": int(to.group(4)),
                })
            if slots:
                result[sheet_name] = slots
    return result


def dropdown_cells_by_sheet():
    """x14拡張データ検証(建材ドロップダウン)のsqrefを展開し、シート名→セルアドレス集合を返す。
    openpyxlは拡張検証を落とすため、worksheet XMLを直接解析する。"""
    # rId(worksheets/sheetN.xml) → シート名 を workbook から得る
    result = {}
    with zipfile.ZipFile(XLSX) as z:
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        relsxml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid_to_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', relsxml))
        name_to_target = {}
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml):
            name, rid = m.group(1), m.group(2)
            if rid in rid_to_target:
                name_to_target[name] = "xl/" + rid_to_target[rid]
        for name, target in name_to_target.items():
            try:
                x = z.read(target).decode("utf-8")
            except KeyError:
                continue
            cells = set()
            for block in re.findall(r'<x14:dataValidation\b.*?</x14:dataValidation>', x, re.S):
                if "部材性能シート" not in block:  # 建材ドロップダウンのみ対象
                    continue
                sqm = re.search(r'<xm:sqref>(.*?)</xm:sqref>', block, re.S)
                if not sqm:
                    continue
                for rng in sqm.group(1).split():
                    try:
                        c1, r1, c2, r2 = range_boundaries(rng)
                        for rr in range(r1, r2 + 1):
                            for cc in range(c1, c2 + 1):
                                cells.add(f"{get_column_letter(cc)}{rr}")
                    except Exception:
                        pass
            if cells:
                result[name] = cells
    return result


def main():
    if not os.path.exists(XLSX):
        print("Excel not found:", XLSX); sys.exit(1)

    dropdowns = dropdown_cells_by_sheet()
    images = images_by_sheet()
    drawing_slots = drawing_slots_by_sheet()

    wb_f = openpyxl.load_workbook(XLSX, data_only=False)  # 数式
    wb_v = openpyxl.load_workbook(XLSX, data_only=True)   # キャッシュ計算値

    model = {"source": os.path.basename(XLSX), "sheetOrder": [], "sheets": {}}
    cached_by_sheet = {}

    for ws in wb_f.worksheets:
        name = ws.title
        wsv = wb_v[name]
        model["sheetOrder"].append(name)
        max_row = ws.max_row
        max_col = ws.max_column

        # dense 2D arrays (HyperFormula buildFromSheets 用)
        data = [[None] * max_col for _ in range(max_row)]
        cached = [[None] * max_col for _ in range(max_row)]
        # スタイル: 一意スタイルテーブル + セルごとのインデックス（サイズ削減）
        style_table = []
        style_index = {}
        styles = [[-1] * max_col for _ in range(max_row)]

        def border_sides(c):
            b = c.border
            return (
                1 if (b.left and b.left.style) else 0,
                1 if (b.right and b.right.style) else 0,
                1 if (b.top and b.top.style) else 0,
                1 if (b.bottom and b.bottom.style) else 0,
            )

        def fill_hex(c):
            f = c.fill
            if f and f.patternType == "solid":
                return resolve_color(f.fgColor, XLSX)
            return None

        def style_id(c):
            al = c.alignment
            ff = c.font
            bd = border_sides(c)
            fillc = fill_hex(c)
            fontc = resolve_color(ff.color, XLSX) if ff and ff.color else None
            key = (
                (al.horizontal or ""),
                (al.vertical or ""),
                bool(ff.bold),
                round(ff.sz or 11, 1),
                (c.number_format or "General"),
                bool(al.wrap_text),
                bd,
                fillc or "",
                fontc or "",
            )
            if key not in style_index:
                style_index[key] = len(style_table)
                style_table.append({
                    "h": key[0], "v": key[1], "b": key[2],
                    "sz": key[3], "fmt": key[4], "wrap": key[5],
                    "bd": list(bd), "fill": fillc, "color": fontc,
                })
            return style_index[key]

        for row in ws.iter_rows():
            for c in row:
                r, col = c.row - 1, c.column - 1
                if r < max_row and col < max_col:
                    data[r][col] = cell_content(c)
                    styles[r][col] = style_id(c)
        for row in wsv.iter_rows():
            for c in row:
                r, col = c.row - 1, c.column - 1
                if r < max_row and col < max_col:
                    cached[r][col] = c.value

        merges = [str(m) for m in ws.merged_cells.ranges]

        # 図面配置枠（EMF枠）にラベルを付与: 枠の上数行・左数列にある◇ラベルを探す
        slots = [dict(s) for s in drawing_slots.get(name, [])]
        for si, s in enumerate(slots):
            label = ""
            for rr in range(max(0, s["fromRow"] - 3), s["fromRow"] + 1):
                for cc in range(s["fromCol"], min(max_col, s["fromCol"] + 6)):
                    v = data[rr][cc] if rr < max_row and cc < max_col else None
                    if isinstance(v, str) and "◇" in v and not v.startswith("="):
                        label = v.strip().lstrip("◇").strip()
                        break
                if label:
                    break
            s["label"] = label or f"図{si + 1}"

        # 入力欄の自動判定: locked=False かつ 数式でない セル（白セル＝ユーザー入力）
        inputs = []
        for row in ws.iter_rows():
            for c in row:
                r, col = c.row - 1, c.column - 1
                if r >= max_row or col >= max_col:
                    continue
                if c.protection and c.protection.locked is False:
                    content = data[r][col]
                    if isinstance(content, str) and content.startswith("="):
                        continue  # 数式は入力ではない
                    inputs.append({
                        "addr": f"{get_column_letter(c.column)}{c.row}",
                        "row": r,
                        "col": col,
                        "default": content,
                    })

        # 列幅(おおよそ): UI再現の参考
        col_widths = {}
        for letter, dim in ws.column_dimensions.items():
            if dim.width:
                col_widths[letter] = round(dim.width, 2)
        row_heights = {}
        for idx, dim in ws.row_dimensions.items():
            if dim.height:
                row_heights[str(idx)] = round(dim.height, 2)

        # 実行時モデル（cachedは含めない＝バンドル軽量化）
        model["sheets"][name] = {
            "maxRow": max_row,
            "maxCol": max_col,
            "data": data,
            "merges": merges,
            "colWidths": col_widths,
            "rowHeights": row_heights,
            "inputs": inputs,
            "dropdownCells": sorted(dropdowns.get(name, [])),
            "styleTable": style_table,
            "styles": styles,
            "images": images.get(name, []),
            "drawingSlots": slots,
            "defaultRowHeight": round(ws.sheet_format.defaultRowHeight or 15, 2),
        }
        cached_by_sheet[name] = cached  # 忠実性テスト専用
        n_formula = sum(1 for r in data for v in r if isinstance(v, str) and v.startswith("="))
        print(f"  {name}: {max_row}x{max_col}, formulas={n_formula}, merges={len(merges)}, inputs={len(inputs)}, images={len(images.get(name, []))}, drawingSlots={len(slots)}")

    os.makedirs(os.path.dirname(OUT_MODEL), exist_ok=True)
    with open(OUT_MODEL, "w", encoding="utf-8") as f:
        json.dump(model, f, ensure_ascii=False)
    print("wrote", OUT_MODEL, f"({os.path.getsize(OUT_MODEL)//1024} KB)")

    # 忠実性テスト専用のキャッシュ値（実行用バンドルには含めない）
    out_cached = os.path.join(ROOT, "src", "data", "workbook-cached.json")
    with open(out_cached, "w", encoding="utf-8") as f:
        json.dump({"sheetOrder": model["sheetOrder"], "cached": cached_by_sheet}, f, ensure_ascii=False)
    print("wrote", out_cached, f"({os.path.getsize(out_cached)//1024} KB)")

    # --- 建材マスタ (B12:G142) ---
    msv = wb_v["部材性能シート"]
    materials = []
    for r in range(12, 143):
        name = msv[f"B{r}"].value
        if name is None or str(name).strip() == "":
            continue
        materials.append({
            "row": r,
            "name": name,
            "R": msv[f"C{r}"].value,           # 熱抵抗値(計算済み)
            "thickness": msv[f"E{r}"].value,   # 厚み mm
            "lambda": msv[f"F{r}"].value,      # 熱伝導率
            "U": msv[f"G{r}"].value,           # 熱貫流率
        })
    with open(OUT_MASTER, "w", encoding="utf-8") as f:
        json.dump({"range": "部材性能シート!B12:B142", "materials": materials},
                  f, ensure_ascii=False, indent=1)
    print("wrote", OUT_MASTER, f"({len(materials)} materials)")

    # --- 画像抽出 (png/jpeg のみ。emf はWeb非対応のため除外) ---
    import zipfile
    os.makedirs(ASSET_DIR, exist_ok=True)
    with zipfile.ZipFile(XLSX) as z:
        imgs = [n for n in z.namelist() if re.match(r"xl/media/.*\.(png|jpe?g)$", n, re.I)]
        for n in imgs:
            dest = os.path.join(ASSET_DIR, os.path.basename(n))
            with z.open(n) as src, open(dest, "wb") as out:
                shutil.copyfileobj(src, out)
        print(f"extracted {len(imgs)} images to public/assets/")


if __name__ == "__main__":
    main()
